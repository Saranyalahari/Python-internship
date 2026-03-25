import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image
import os

# -----------------------------
# FUNCTIONS
# -----------------------------
def load_file():
    global df
    file_path = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])
    if file_path:
        df = pd.read_csv(file_path)
        messagebox.showinfo("Success", "File Loaded Successfully!")

def generate_report():
    if df is None:
        messagebox.showerror("Error", "Please load a CSV file first!")
        return

    # -----------------------------
    # Pivot Table
    # -----------------------------
    pivot = pd.pivot_table(df, values='Sales', index='Category', aggfunc='sum')

    # -----------------------------
    # Summary Stats
    # -----------------------------
    total_sales = df['Sales'].sum()
    avg_sales = df['Sales'].mean()
    max_sales = df['Sales'].max()

    # -----------------------------
    # Chart
    # -----------------------------
    chart_path = "chart.png"
    pivot.plot(kind='bar', legend=False)
    plt.title("Sales by Category")
    plt.ylabel("Sales")
    plt.tight_layout()
    plt.savefig(chart_path)
    plt.close()

    # -----------------------------
    # Excel Report
    # -----------------------------
    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    # Title
    ws["A1"] = "Sales Report"
    ws["A1"].font = Font(size=14, bold=True)

    # Pivot Table
    ws.append(["Category", "Total Sales"])
    for row in pivot.itertuples():
        ws.append([row.Index, row.Sales])

    # Summary
    ws.append([])
    ws.append(["Summary"])
    ws.append(["Total Sales", total_sales])
    ws.append(["Average Sales", avg_sales])
    ws.append(["Max Sales", max_sales])

    # Insert Chart Image
    img = Image(chart_path)
    ws.add_image(img, "E2")

    wb.save(save_path)

    messagebox.showinfo("Success", "Excel Report Generated!")

# -----------------------------
# GUI
# -----------------------------
df = None

root = tk.Tk()
root.title("Excel Report Generator")
root.geometry("400x250")

tk.Label(root, text="📊 Excel Report Generator", font=("Arial", 14)).pack(pady=10)

tk.Button(root, text="Load CSV", command=load_file, width=20).pack(pady=10)
tk.Button(root, text="Generate Report", command=generate_report, width=20).pack(pady=10)

root.mainloop()